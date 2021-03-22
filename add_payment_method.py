from selenium import webdriver
import xlrd
import xlsxwriter
import openpyxl
from xlwt import Workbook
# import
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from pynput.keyboard import Key, Controller

keyboard = Controller()
loc="C:\\Users\Arsal Azeem\\Desktop\\paymentcase.xlsx"
wb=xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
wb2 = openpyxl.open(loc)
sheet2 = wb2.active
k=1
totalcount=sheet.nrows
driver = webdriver.Chrome(
        executable_path='C:\\Users\\Arsal Azeem\\Desktop\\drivers\\chromedriver_win32\\chromedriver.exe')
driver.get("https://dev.tradishes.com/#/register")
driver.implicitly_wait(40)
driver.find_element(By.XPATH, '//*[@id="loginForm"]/div[2]/label').click()
keyboard.press(Key.tab)
keyboard.release(Key.tab)
keyboard.type("arsalazeem101@gmail.com")
keyboard.press(Key.tab)
keyboard.release(Key.tab)
keyboard.type("12345678")
driver.find_element(By.XPATH, '//*[@id="submitLoginForm"]').click()
driver.find_element(By.XPATH,'//*[@id="dropdownMenuButton"]').click()
driver.find_element(By.XPATH,'//*[@id="nav-menu-container"]/ul/li[2]/div/div/a[2]').click()
driver.find_element(By.XPATH,'//*[@id="payment-customer"]/div/div[1]/div/div/div[2]/div').click()
for i in range(0,sheet.nrows-1):
        driver.find_element(By.XPATH, '//*[@id="card-name-element"]').clear()
        driver.find_element(By.XPATH, '//*[@id="card_number_input"]').clear()
        driver.find_element(By.XPATH, '//*[@id="card_exp_input"]').clear()
        driver.find_element(By.XPATH, '//*[@id="card_exp_cvc"]').clear()
        driver.find_element(By.XPATH, '//*[@id="card_exp_zip"]').clear()
        name= str(sheet.cell_value(k, 0))
        cardnumber= str(sheet.cell_value(k, 1))
        expiry_date= str(sheet.cell_value(k, 2))
        cvc = str(int(sheet.cell_value(k, 3)))
        zip_code = str(int(sheet.cell_value(k, 4)))
        try:
                driver.find_element(By.XPATH,'//*[@id="card-name-element"]').send_keys(name)
                driver.find_element(By.XPATH,'//*[@id="card_number_input"]').send_keys(cardnumber)
                driver.find_element(By.XPATH,'//*[@id="card_exp_input"]').send_keys(expiry_date)
                driver.find_element(By.XPATH,'//*[@id="card_exp_cvc"]').send_keys(cvc)
                driver.find_element(By.XPATH,'//*[@id="card_exp_zip"]').send_keys(zip_code)
                driver.find_element(By.XPATH,'//*[@id="payment-form"]/div[5]/button[2]').click()
                element=driver.find_element(By.XPATH,'//*[@id="payment-form"]/div[4]/div[2]')
                val=element.is_displayed()
                if val==True:
                        n = k
                        n = n + 1
                        c1 = sheet2.cell(row=n, column=6)
                        c1.value = "Failed"
                        wb2.save(loc)

                else:
                        n = k
                        n = n + 1
                        c1 = sheet2.cell(row=n, column=6)
                        c1.value = "Pass"
                        wb2.save(loc)

        except:
                n = k
                n = n + 1
                c1 = sheet2.cell(row=n, column=8)
                c1.value = "Failed"
                wb2.save(loc)
        k=k+1

        
time.sleep(5)

