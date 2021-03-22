from selenium import webdriver
import xlrd
# import xlsxwriter
import pdb
import openpyxl
import stringfy
import xlsxwriter
import requests
import random
import json
import logging

import json
import initchrome
from xlwt import Workbook
# import
import generate_sample_values_excel
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from datetime import date
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import kardkeymakeorder
from pynput.keyboard import Key, Controller
import init_excel
loc= "C:\\Users\\Enabling\\Desktop\\Enabling\\data\\data2.xlsx"

def populate_excel(limit):
    try:
        path = loc

        temp = init_excel.init_excel(path)
        sheet = temp[0]
        wb = temp[1]
        email = sheet.cell(row=1, column=1)
        email.value = "Email"

        first_name = sheet.cell(row=1, column=2)
        first_name.value = "First Name"

        last_name = sheet.cell(row=1, column=3)
        last_name.value = "Last Name"

        adress1 = sheet.cell(row=1, column=4)
        adress1.value = "Adress1"

        adress2 = sheet.cell(row=1, column=5)
        adress2.value = "Adress2"

        city = sheet.cell(row=1, column=6)
        city.value = "city"

        zip_code = sheet.cell(row=1, column=7)
        zip_code.value = "zip code"

        phone_number = sheet.cell(row=1, column=8)
        phone_number.value = "phone_number"

        card_number = sheet.cell(row=1, column=9)
        card_number.value = "card_number"

        card_name = sheet.cell(row=1, column=10)
        card_name.value = "card_name"

        Expiry_Date = sheet.cell(row=1, column=11)
        Expiry_Date.value = "Expiry_Date"

        no_of_keys = sheet.cell(row=1, column=12)
        no_of_keys.value = "No_of_Keys"

        actual_result = sheet.cell(row=1, column=13)
        actual_result.value = "Actual_Result"

        expected_result = sheet.cell(row=1, column=14)
        expected_result.value = "Expected_Result"
        for i in range(2, limit):
            decision=random.randint(1,2)
            if decision==1:
                email = sheet.cell(row=i, column=1)
                email.value = generate_sample_values_excel.generate_random_email()

                first_name = sheet.cell(row=i, column=2)
                first_name.value = generate_sample_values_excel.generate_random_first_name()

                last_name = sheet.cell(row=i, column=3)
                last_name.value = generate_sample_values_excel.generate_random_last_name()

                adress1 = sheet.cell(row=i, column=4)
                adress1.value = generate_sample_values_excel.generate_random_adress()

                adress2 = sheet.cell(row=i, column=5)
                adress2.value = generate_sample_values_excel.generate_random_adress()

                city = sheet.cell(row=i, column=6)
                city.value = generate_sample_values_excel.generate_random_city()

                zip_code = sheet.cell(row=i, column=7)
                zip_code.value = generate_sample_values_excel.generate_random_zipcodes()

                phone_number = sheet.cell(row=i, column=8)
                phone_number.value = str(generate_sample_values_excel.generate_phone_number())

                card_number = sheet.cell(row=i, column=9)
                card_number.value = generate_sample_values_excel.generate_random_card_number()

                card_name = sheet.cell(row=i, column=10)
                card_name.value = generate_sample_values_excel.generate_random_name()

                expiry_date = sheet.cell(row=i, column=11)
                expiry_date.value = generate_sample_values_excel.generate_random_date()

                no_of_keys = sheet.cell(row=i, column=12)
                no_of_keys.value = generate_sample_values_excel.generate_random_quantity(20)

                actual_result = sheet.cell(row=i, column=13)
                actual_result.value = "No Result"

                expected_result = sheet.cell(row=i, column=14)
                expected_result.value = "Pass"
            else:
                email = sheet.cell(row=i, column=1)
                email.value = generate_sample_values_excel.generate_bad_email()

                first_name = sheet.cell(row=i, column=2)
                first_name.value = generate_sample_values_excel.generate_random_first_name()

                last_name = sheet.cell(row=i, column=3)
                last_name.value = generate_sample_values_excel.generate_random_last_name()

                adress1 = sheet.cell(row=i, column=4)
                adress1.value = generate_sample_values_excel.generate_random_adress()

                adress2 = sheet.cell(row=i, column=5)
                adress2.value = generate_sample_values_excel.generate_random_adress()

                city = sheet.cell(row=i, column=6)
                city.value = generate_sample_values_excel.generate_bad_city()

                zip_code = sheet.cell(row=i, column=7)
                zip_code.value = generate_sample_values_excel.generate_bad_zip_code()

                phone_number = sheet.cell(row=i, column=8)
                phone_number.value = str(generate_sample_values_excel.generate_bad_phone_number())

                card_number = sheet.cell(row=i, column=9)
                card_number.value = generate_sample_values_excel.generate_random_card_number()

                card_name = sheet.cell(row=i, column=10)
                card_name.value = generate_sample_values_excel.generate_random_name()

                expiry_date = sheet.cell(row=i, column=11)
                expiry_date.value = generate_sample_values_excel.generate_bad_date()

                no_of_keys = sheet.cell(row=i, column=12)
                no_of_keys.value = generate_sample_values_excel.generate_random_quantity(20)

                actual_result = sheet.cell(row=i, column=13)
                actual_result.value = "No Result"

                expected_result = sheet.cell(row=i, column=14)
                expected_result.value = "Fail"


        wb.save(path)
    except Exception as e:
        # wb.close()
        # populate_excel()
        print(e)



def clear_excel_file():
    file = open(loc, "r+")
    file.truncate(0)
    file.close()

try:
    populate_excel(30)
except Exception as e:
    print(e)
    workbook = xlsxwriter.Workbook(loc)
    populate_excel()
    

# populate_excel(9999)

